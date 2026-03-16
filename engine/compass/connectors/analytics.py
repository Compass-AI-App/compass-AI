"""Analytics connector — the DATA source of truth.

Ingests: CSV/JSON/Excel usage data, metrics exports, analytics summaries.
Answers: "What IS happening?"

Enhanced with: structure detection, auto-detect column types, time series
vs categorical classification, and Excel (.xlsx/.xls) support.
"""

from __future__ import annotations

import csv
import json
import logging
import re
from pathlib import Path

from compass.connectors.base import Connector
from compass.models.sources import Evidence, SourceType

logger = logging.getLogger(__name__)

MAX_ROWS = 500


def _detect_column_type(values: list[str]) -> str:
    """Auto-detect column type from sample values.

    Returns one of: 'numeric', 'date', 'boolean', 'categorical'.
    """
    if not values:
        return "categorical"

    # Sample non-empty values
    sample = [v.strip() for v in values if v and v.strip()][:50]
    if not sample:
        return "categorical"

    # Check numeric
    numeric_count = 0
    for v in sample:
        cleaned = v.replace(",", "").replace("$", "").replace("%", "").strip()
        try:
            float(cleaned)
            numeric_count += 1
        except ValueError:
            pass
    if numeric_count / len(sample) > 0.8:
        return "numeric"

    # Check date patterns
    date_patterns = [
        r"\d{4}-\d{2}-\d{2}",
        r"\d{1,2}/\d{1,2}/\d{2,4}",
        r"\d{1,2}-\w{3}-\d{2,4}",
    ]
    date_count = sum(
        1 for v in sample if any(re.match(p, v) for p in date_patterns)
    )
    if date_count / len(sample) > 0.8:
        return "date"

    # Check boolean
    bool_values = {"true", "false", "yes", "no", "1", "0", "y", "n"}
    bool_count = sum(1 for v in sample if v.lower() in bool_values)
    if bool_count / len(sample) > 0.8:
        return "boolean"

    return "categorical"


def _classify_dataset(headers: list[str], column_types: dict[str, str]) -> str:
    """Classify dataset as time_series, categorical, or mixed."""
    has_date = any(t == "date" for t in column_types.values())
    has_numeric = any(t == "numeric" for t in column_types.values())

    if has_date and has_numeric:
        return "time_series"
    if has_numeric and not has_date:
        return "categorical"
    return "mixed"


def _compute_stats(values: list[str], col_type: str) -> dict:
    """Compute basic stats for a column."""
    if col_type == "numeric":
        nums = []
        for v in values:
            cleaned = v.replace(",", "").replace("$", "").replace("%", "").strip()
            try:
                nums.append(float(cleaned))
            except ValueError:
                pass
        if nums:
            return {
                "min": round(min(nums), 2),
                "max": round(max(nums), 2),
                "mean": round(sum(nums) / len(nums), 2),
                "count": len(nums),
            }
    elif col_type == "categorical":
        counts: dict[str, int] = {}
        for v in values:
            v = v.strip()
            if v:
                counts[v] = counts.get(v, 0) + 1
        top = sorted(counts.items(), key=lambda x: -x[1])[:10]
        return {"unique": len(counts), "top": dict(top)}
    elif col_type == "date":
        non_empty = [v.strip() for v in values if v.strip()]
        if non_empty:
            return {"earliest": min(non_empty), "latest": max(non_empty), "count": len(non_empty)}

    return {}


class AnalyticsConnector(Connector):
    """Ingests analytics and usage data with structure detection."""

    connector_type = "analytics"

    def validate(self) -> bool:
        path = self.config.path
        if not path:
            return False
        return Path(path).expanduser().exists()

    def ingest(self) -> list[Evidence]:
        path = self.config.path
        if not path:
            return []

        p = Path(path).expanduser().resolve()
        evidence: list[Evidence] = []

        if p.is_file():
            ev = self._ingest_file(p)
            if ev:
                evidence.extend(ev)
        elif p.is_dir():
            for fpath in sorted(p.rglob("*")):
                if fpath.is_file() and fpath.suffix.lower() in {".csv", ".json", ".jsonl", ".xlsx", ".xls"}:
                    ev = self._ingest_file(fpath)
                    if ev:
                        evidence.extend(ev)

        return evidence

    def _ingest_file(self, fpath: Path) -> list[Evidence]:
        suffix = fpath.suffix.lower()
        if suffix == ".csv":
            return self._ingest_csv(fpath)
        elif suffix in {".json", ".jsonl"}:
            return self._ingest_json(fpath)
        elif suffix in {".xlsx", ".xls"}:
            return self._ingest_excel(fpath)
        return []

    def _detect_delimiter(self, fpath: Path) -> str:
        """Detect CSV delimiter by sniffing the first few lines."""
        try:
            with open(fpath, newline="", errors="ignore") as f:
                sample = f.read(4096)
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            return dialect.delimiter
        except csv.Error:
            return ","

    def _ingest_csv(self, fpath: Path) -> list[Evidence]:
        try:
            delimiter = self._detect_delimiter(fpath)
            with open(fpath, newline="", errors="ignore") as f:
                first_char = f.read(1)
                if first_char != "\ufeff":
                    f.seek(0)
                reader = csv.DictReader(f, delimiter=delimiter)
                rows = []
                for i, row in enumerate(reader):
                    if i >= MAX_ROWS:
                        break
                    rows.append(row)

            if not rows:
                return []

            return self._build_evidence_from_rows(rows, fpath, "csv")
        except Exception as e:
            logger.warning("Failed to ingest CSV %s: %s", fpath, e)
            return []

    def _ingest_excel(self, fpath: Path) -> list[Evidence]:
        """Ingest Excel files using openpyxl."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.warning("openpyxl not installed, skipping Excel file %s", fpath)
            return []

        try:
            wb = load_workbook(fpath, read_only=True, data_only=True)
            evidence: list[Evidence] = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_data: list[dict] = []
                headers: list[str] = []

                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i == 0:
                        headers = [str(h) if h is not None else f"col_{j}" for j, h in enumerate(row)]
                        continue
                    if i > MAX_ROWS:
                        break
                    row_dict = {}
                    for j, val in enumerate(row):
                        if j < len(headers):
                            row_dict[headers[j]] = str(val) if val is not None else ""
                    rows_data.append(row_dict)

                if rows_data:
                    ev = self._build_evidence_from_rows(
                        rows_data, fpath, "excel",
                        extra_meta={"sheet": sheet_name}
                    )
                    evidence.extend(ev)

            wb.close()
            return evidence
        except Exception as e:
            logger.warning("Failed to ingest Excel %s: %s", fpath, e)
            return []

    def _build_evidence_from_rows(
        self,
        rows: list[dict],
        fpath: Path,
        file_type: str,
        extra_meta: dict | None = None,
    ) -> list[Evidence]:
        """Build evidence with structure detection and column typing."""
        headers = list(rows[0].keys())

        # Detect column types
        column_types: dict[str, str] = {}
        column_stats: dict[str, dict] = {}
        for header in headers:
            values = [row.get(header, "") for row in rows]
            col_type = _detect_column_type(values)
            column_types[header] = col_type
            column_stats[header] = _compute_stats(values, col_type)

        # Classify dataset
        dataset_type = _classify_dataset(headers, column_types)

        # Build structured summary
        summary_lines = [
            f"Dataset: {fpath.name}",
            f"Type: {dataset_type}",
            f"Columns: {len(headers)} | Rows: {len(rows)}",
            "",
            "**Column Analysis:**",
        ]

        for header in headers:
            ct = column_types[header]
            stats = column_stats[header]
            stats_str = ""
            if ct == "numeric" and stats:
                stats_str = f" (min: {stats['min']}, max: {stats['max']}, mean: {stats['mean']})"
            elif ct == "categorical" and stats:
                top_items = list(stats.get("top", {}).items())[:5]
                if top_items:
                    top_str = ", ".join(f"{k}: {v}" for k, v in top_items)
                    stats_str = f" ({stats['unique']} unique — top: {top_str})"
            elif ct == "date" and stats:
                stats_str = f" ({stats.get('earliest', '')} to {stats.get('latest', '')})"
            summary_lines.append(f"- {header} [{ct}]{stats_str}")

        summary_lines.append("")
        summary_lines.append("**Sample data:**")

        for row in rows[:15]:
            summary_lines.append(" | ".join(f"{k}: {v}" for k, v in row.items()))

        if len(rows) > 15:
            summary_lines.append(f"... and {len(rows) - 15} more rows")

        metadata = {
            "file": str(fpath),
            "type": file_type,
            "rows": len(rows),
            "columns": headers,
            "column_types": column_types,
            "dataset_type": dataset_type,
        }
        if extra_meta:
            metadata.update(extra_meta)

        return [Evidence(
            source_type=SourceType.DATA,
            connector="analytics",
            title=f"Analytics: {fpath.stem} ({dataset_type}, {len(rows)} rows)",
            content="\n".join(summary_lines),
            metadata=metadata,
        )]

    def _ingest_json(self, fpath: Path) -> list[Evidence]:
        try:
            content = fpath.read_text(errors="ignore")
            if fpath.suffix.lower() == ".jsonl":
                data = []
                for line in content.splitlines():
                    line = line.strip()
                    if line:
                        data.append(json.loads(line))
            else:
                data = json.loads(content)

            if isinstance(data, list) and data and isinstance(data[0], dict):
                # Treat as tabular data — convert to rows
                rows = [{str(k): str(v) for k, v in item.items()} for item in data[:MAX_ROWS]]
                return self._build_evidence_from_rows(rows, fpath, "json")

            # Fallback for non-tabular JSON
            if isinstance(data, list):
                summary = json.dumps(data[:20], indent=2)
                count = len(data)
            elif isinstance(data, dict):
                summary = json.dumps(data, indent=2)[:5000]
                count = len(data)
            else:
                summary = str(data)[:5000]
                count = 1

            return [Evidence(
                source_type=SourceType.DATA,
                connector="analytics",
                title=f"Analytics: {fpath.stem}",
                content=f"Dataset: {fpath.name}\nRecords: {count}\n\n{summary}",
                metadata={"file": str(fpath), "type": "json", "records": count},
            )]
        except Exception as e:
            logger.warning("Failed to ingest JSON %s: %s", fpath, e)
            return []
